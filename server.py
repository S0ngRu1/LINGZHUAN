import os
import logging
import re
from datetime import datetime, timedelta
from typing import Optional, List, Dict
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import pandas as pd
import io

# --- 新增的库导入 ---
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from app.src.llm_parser import parse_schedule_from_text
from app.utils.file_handler import read_file_content

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 初始化Flask应用
app = Flask(__name__)
CORS(app)  # 允许跨域请求

# 设置一个用于临时存放上传文件的目录
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def _parse_date_range(text: str) -> tuple:
    if not isinstance(text, str):
        return (None, None)
    start_str, end_str = None, None
    range_match = re.search(r'[（\(](\d{1,2}\.\d{1,2})-(\d{1,2}\.\d{1,2})[）\)]', text)
    if range_match:
        start_str, end_str = range_match.groups()
    else:
        single_match = re.search(r'[（\(](\d{1,2}\.\d{1,2})[）\)]', text)
        if single_match:
            date_str = single_match.group(1)
            start_str = date_str
            end_str = date_str
        else:
            try:
                single_date = pd.to_datetime(text, errors='coerce')
                if pd.notna(single_date):
                    date_str_formatted = single_date.strftime('%Y-%m-%d')
                    return (date_str_formatted, date_str_formatted)
            except Exception:
                pass
            return (None, None)
    if start_str is None:
        return (None, None)
    current_year = datetime.now().year
    try:
        start_month, start_day = map(int, start_str.split('.'))
        end_month, end_day = map(int, end_str.split('.'))
        start_date = datetime(current_year, start_month, start_day)
        end_year = current_year + 1 if end_month < start_month else current_year
        end_date = datetime(end_year, end_month, end_day)
        return (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
    except (ValueError, AttributeError):
        logger.warning(f"在文本 '{text}' 中发现无效日期，已忽略。")
        return (None, None)


def parse_excel_to_records(file_path: str) -> Optional[List[Dict]]:
    try:
        df = pd.read_excel(file_path, sheet_name=0, header=1)
        column_mapping = {
            "工作步骤": "阶段", "工作内容": "工作子项", "成果文件": "成果文件",
            "主责单位": "主责单位/负责人", "完成时限": "完成时限"
        }
        required_input_cols = list(column_mapping.keys())
        if not all(col in df.columns for col in required_input_cols):
            missing_cols = [col for col in required_input_cols if col not in df.columns]
            logger.error(f"Excel文件中缺少必需的列: {', '.join(missing_cols)}")
            return None
        df.rename(columns=column_mapping, inplace=True)
        df[['开始日期', '结束日期']] = df['完成时限'].apply(lambda x: pd.Series(_parse_date_range(x)))
        target_columns = ["阶段", "工作子项", "成果文件", "主责单位/负责人", "开始日期", "结束日期"]
        df = df[target_columns]
        columns_to_fill = ["阶段", "主责单位/负责人"]
        df[columns_to_fill] = df[columns_to_fill].ffill()
        df = df.where(pd.notna(df), None)
        records = df.to_dict('records')
        return records
    except Exception as e:
        logger.error(f"直接解析Excel文件失败: {e}")
        return None


def generate_gantt_and_charts_excel(work_data: List[Dict]):
    """
    根据工作数据生成一个包含甘特图、原始数据和统计图表的Excel文件。
    文件在内存中生成，直接返回二进制数据。
    """
    df = pd.DataFrame(work_data)

    # 1. 数据预处理，将日期字符串转换为datetime对象，便于后续计算
    df['start_date_dt'] = pd.to_datetime(df['开始日期'], errors='coerce')
    df['end_date_dt'] = pd.to_datetime(df['结束日期'], errors='coerce')

    # 2. 数据聚合统计 (用于图表)
    phase_counts = df['阶段'].value_counts().reset_index()
    phase_counts.columns = ['阶段', '任务数量']
    person_counts = df['主责单位/负责人'].value_counts().reset_index()
    person_counts.columns = ['主责单位/负责人', '任务数量']

    # 3. 创建Excel工作簿和工作表
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    gantt_sheet = wb.create_sheet(title="项目进度甘特图")
    analysis_sheet = wb.create_sheet(title="统计分析")
    data_sheet = wb.create_sheet(title="原始数据")
    wb.move_sheet(analysis_sheet, offset=-len(wb.sheetnames))  # 将统计分析页置顶

    # 4. 填充“统计分析”工作表并创建图表
    analysis_sheet['A1'] = '各阶段任务数量统计'
    analysis_sheet['A1'].font = Font(bold=True, size=14)
    for r in dataframe_to_rows(phase_counts, index=False, header=True):
        analysis_sheet.append(r)

    analysis_sheet['D1'] = '各主责单位/负责人任务分布统计'
    analysis_sheet['D1'].font = Font(bold=True, size=14)
    analysis_sheet['D2'] = '主责单位/负责人'
    analysis_sheet['D2'].font = Font(bold=True)
    analysis_sheet['E2'] = '任务数量'
    analysis_sheet['E2'].font = Font(bold=True)
    for index, row_data in person_counts.iterrows():
        row_num = index + 3  # +3因为iterrows索引从0开始，而我们的数据从第3行开始
        analysis_sheet.cell(row=row_num, column=4, value=row_data['主责单位/负责人'])
        analysis_sheet.cell(row=row_num, column=5, value=row_data['任务数量'])

    # 创建柱状图 (阶段任务数)
    bar_chart = BarChart()
    bar_chart.title = "各阶段任务数量"
    data = Reference(analysis_sheet, min_col=2, min_row=2, max_row=len(phase_counts) + 1)
    cats = Reference(analysis_sheet, min_col=1, min_row=3, max_row=len(phase_counts) + 1)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    analysis_sheet.add_chart(bar_chart, "A12")

    # 创建饼图 (负责人任务分布)
    pie_chart = PieChart()
    pie_chart.title = "各主责单位/负责人任务分布"
    labels = Reference(analysis_sheet, min_col=4, min_row=3, max_row=len(person_counts) + 2)
    # 数据从E2(包含表头)到数据末尾
    data = Reference(analysis_sheet, min_col=5, min_row=2, max_row=len(person_counts) + 2)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    analysis_sheet.add_chart(pie_chart, "H12")

    # 5. 填充“原始数据”工作表
    # 从DataFrame中移除临时的datetime列
    df_for_sheet = df.drop(columns=['start_date_dt', 'end_date_dt'])
    for r in dataframe_to_rows(df_for_sheet, index=False, header=True):
        data_sheet.append(r)

    # 6. 绘制甘特图
    # 颜色定义
    colors = ['B7DEE8', 'FCD5B4', 'C4D79B', 'D8BFD8', 'F0C2C2', 'FFFACD']
    light_colors = ['E8F5F8', 'FEF3E6', 'EFF5DA', 'F4ECF7', 'FBE9E7', 'FFFDE7']
    unique_phases = df['阶段'].unique()
    phase_color_map = {phase: {'task': colors[i % len(colors)], 'info': light_colors[i % len(light_colors)]} for
                       i, phase in enumerate(unique_phases)}

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # 设置表头
    headers = ["阶段", "工作子项", "成果文件", "主责单位/负责人"]
    for i, header in enumerate(headers):
        gantt_sheet.cell(row=1, column=i + 1, value=header).font = header_font

    # 计算日历范围
    valid_dates = df.dropna(subset=['start_date_dt', 'end_date_dt'])
    if not valid_dates.empty:
        min_date = valid_dates['start_date_dt'].min()
        max_date = valid_dates['end_date_dt'].max()
        date_range = [min_date + timedelta(days=d) for d in range((max_date - min_date).days + 1)]

        for i, date in enumerate(date_range):
            cell = gantt_sheet.cell(row=1, column=len(headers) + 1 + i, value=date)
            cell.font = header_font
            cell.alignment = center_align
            cell.number_format = 'MM-DD'

    # 填充任务数据和颜色
    for r_idx, task in df.iterrows():
        row_num = r_idx + 2
        colorset = phase_color_map.get(task['阶段'], {'info': 'FFFFFF', 'task': 'D9D9D9'})
        info_fill = PatternFill(start_color=colorset['info'], end_color=colorset['info'], fill_type="solid")
        task_fill = PatternFill(start_color=colorset['task'], end_color=colorset['task'], fill_type="solid")

        gantt_sheet.cell(row=row_num, column=1, value=task['阶段']).fill = info_fill
        gantt_sheet.cell(row=row_num, column=2, value=task['工作子项']).fill = info_fill
        gantt_sheet.cell(row=row_num, column=3, value=task['成果文件']).fill = info_fill
        gantt_sheet.cell(row=row_num, column=4, value=task['主责单位/负责人']).fill = info_fill

        if pd.notna(task['start_date_dt']) and pd.notna(task['end_date_dt']):
            for d_idx, cal_date in enumerate(date_range):
                if task['start_date_dt'] <= cal_date <= task['end_date_dt']:
                    gantt_sheet.cell(row=row_num, column=len(headers) + 1 + d_idx).fill = task_fill

    # 设置列宽
    gantt_sheet.column_dimensions['A'].width = 15
    gantt_sheet.column_dimensions['B'].width = 30
    gantt_sheet.column_dimensions['C'].width = 25
    gantt_sheet.column_dimensions['D'].width = 20

    # 7. 将工作簿保存在内存中
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/api/parse-file', methods=['POST'])
def parse_file():
    if 'file' not in request.files:
        return jsonify({"error": "请求中未找到文件部分"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "未选择任何文件"}), 400
    if file:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        _, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()
        file.save(filepath)
        try:
            parsed_data = None
            if file_extension in ['.xlsx', '.xls']:
                logger.info("检测到Excel文件，进行直接解析...")
                parsed_data = parse_excel_to_records(filepath)
                if parsed_data is None:
                    return jsonify({"error": "无法直接解析此Excel文件，请检查文件格式和内容。"}), 500
            else:
                logger.info(f"检测到 {file_extension} 文件，使用LLM进行解析...")
                content = read_file_content(filepath)
                if content is None:
                    return jsonify({"error": "不支持的文件类型或文件读取失败"}), 400

                parsed_data = parse_schedule_from_text(content)

            os.remove(filepath)
            if parsed_data:
                logger.info("解析成功，返回数据。")
                return jsonify(parsed_data)
            else:
                logger.error("解析失败或未返回有效数据。")
                return jsonify({"error": "内容解析失败，请检查文件内容或API配置"}), 500

        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            logger.error(f"处理过程中发生错误: {e}")
            return jsonify({"error": f"服务器内部错误: {e}"}), 500
    return jsonify({"error": "未知错误"}), 500


# --- 新增API接口，用于生成和下载Excel文件 ---
@app.route('/api/generate-excel', methods=['POST'])
def generate_excel_endpoint():
    """
    接收前端发来的JSON格式任务数据，生成包含图表的Excel文件并返回给用户下载。
    """
    # 1. 从请求体中获取JSON数据
    work_data = request.get_json()

    # 2. 验证数据有效性
    if not isinstance(work_data, list) or not work_data:
        logger.error("接收到的生成请求数据无效或为空")
        return jsonify({"error": "无效或空的数据"}), 400

    try:
        # 3. 调用核心函数在内存中生成Excel文件
        logger.info("开始生成包含图表的Excel文件...")
        excel_buffer = generate_gantt_and_charts_excel(work_data)
        logger.info("文件生成成功，准备发送给客户端。")

        # 4. 使用send_file将内存中的文件作为附件发送给前端
        filename = f"项目计划甘特图_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"生成Excel文件时发生严重错误: {e}", exc_info=True)
        return jsonify({"error": "服务器在生成Excel文件时发生内部错误"}), 500


if __name__ == '__main__':
    # 在实际部署时，请关闭debug模式
    app.run(host='0.0.0.0', port=5000, debug=True)