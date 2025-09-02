from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import pandas as pd
import os
from config.excel_config import (
    EXCEL_SHEET_NAMES, EXCEL_COLUMN_WIDTH, EXCEL_EXPORT_PATH, EXCEL_HEADERS
)
from utils.file_utils import delete_file_if_exist, create_dir_if_not_exist
from utils.log_utils import logger

def export_to_excel(processed_data: pd.DataFrame, gantt_image_path: str) -> str:
    """
    将处理后的数据和甘特图图片导出到Excel
    :param processed_data: 处理后的工作数据
    :param gantt_image_path: 甘特图图片路径
    :return: Excel文件保存路径
    """
    try:
        # 确保导出目录存在
        export_dir = os.path.dirname(EXCEL_EXPORT_PATH)
        create_dir_if_not_exist(export_dir)
        
        # 先删除旧的Excel文件（避免版本冲突）
        delete_file_if_exist(EXCEL_EXPORT_PATH)
        
        # 1. 初始化Excel工作簿
        wb = Workbook()
        
        # 2. 第一个工作表：甘特图数据明细
        ws_data = wb.active
        ws_data.title = EXCEL_SHEET_NAMES["data_detail"]
        
        # 写入表头（加粗样式）
        header_font = Font(bold=True)
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
        
        # 写入数据（按表头顺序匹配字段）
        data_fields = [
            "阶段", "工作子项", "开始日", "结束日", "持续天数", "关键产出物", "阶段颜色（十六进制）"
        ]
        for row_idx, (_, data_row) in enumerate(processed_data.iterrows(), 2):
            for col_idx, field in enumerate(data_fields, 1):
                # 处理数字类型，确保正确显示（避免科学计数法）
                value = data_row[field]
                if isinstance(value, (int, float)):
                    ws_data.cell(row=row_idx, column=col_idx).number_format = '0'
                ws_data.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽（按配置文件设置）
        for col, width in EXCEL_COLUMN_WIDTH.items():
            ws_data.column_dimensions[col].width = width
        
        # 3. 第二个工作表：甘特图可视化（插入图片）
        ws_image = wb.create_sheet(title=EXCEL_SHEET_NAMES["gantt_image"])
        
        # 验证图片是否存在
        if not os.path.exists(gantt_image_path):
            error_msg = f"甘特图图片不存在：{gantt_image_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        # 插入甘特图图片并调整大小
        img = Image(gantt_image_path)
        # 按比例调整图片大小（保持宽高比）
        img.width = 1200  # 宽度调整为1200像素
        img.height = int(img.height * (1200 / img.width))  # 按比例计算高度
        ws_image.add_image(img, 'A1')  # 从A1单元格开始插入
        
        # 4. 保存Excel文件
        wb.save(EXCEL_EXPORT_PATH)
        logger.info(f"Excel文件导出成功，保存路径：{EXCEL_EXPORT_PATH}")
        
        return EXCEL_EXPORT_PATH
    
    except Exception as e:
        logger.error(f"Excel导出失败：{str(e)}", exc_info=True)
        raise  # 向上抛出异常，让调用者处理
