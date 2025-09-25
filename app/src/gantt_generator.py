# -*- coding: utf-8 -*-
# @Time : 2025/9/5 17:48
# @File : gantt_generator.py

"""
甘特图生成代码
"""


from typing import List, Dict
import pandas as pd
import io

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def generate_gantt_and_charts_excel(work_data: List[Dict]):
    """
    根据工作数据生成一个包含甘特图、原始数据和统计图表的Excel文件。
    文件在内存中生成，直接返回二进制数据。
    """
    df = pd.DataFrame(work_data)

    # 1. 数据预处理，将日期字符串转换为datetime对象，便于后续计算
    df['start_date_dt'] = pd.to_datetime(df['开始日期'], errors='coerce')
    df['end_date_dt'] = pd.to_datetime(df['结束日期'], errors='coerce')

    # 2. 数据聚合统计 (用于图表)
    phase_counts = df['阶段'].value_counts().reset_index()
    phase_counts.columns = ['阶段', '任务数量']
    person_counts = df['主责单位/负责人'].value_counts().reset_index()
    person_counts.columns = ['主责单位/负责人', '任务数量']

    # 3. 创建Excel工作簿和工作表
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    gantt_sheet = wb.create_sheet(title="项目进度甘特图")
    analysis_sheet = wb.create_sheet(title="统计分析")
    data_sheet = wb.create_sheet(title="原始数据")
    wb.move_sheet(analysis_sheet, offset=-len(wb.sheetnames))  # 将统计分析页置顶

    # 4. 填充“统计分析”工作表并创建图表
    analysis_sheet['A1'] = '各阶段任务数量统计'
    analysis_sheet['A1'].font = Font(bold=True, size=14)
    for r in dataframe_to_rows(phase_counts, index=False, header=True):
        analysis_sheet.append(r)

    analysis_sheet['D1'] = '各主责单位/负责人任务分布统计'
    analysis_sheet['D1'].font = Font(bold=True, size=14)
    analysis_sheet['D2'] = '主责单位/负责人'
    analysis_sheet['D2'].font = Font(bold=True)
    analysis_sheet['E2'] = '任务数量'
    analysis_sheet['E2'].font = Font(bold=True)
    for index, row_data in person_counts.iterrows():
        row_num = index + 3  # +3因为iterrows索引从0开始，而我们的数据从第3行开始
        analysis_sheet.cell(row=row_num, column=4, value=row_data['主责单位/负责人'])
        analysis_sheet.cell(row=row_num, column=5, value=row_data['任务数量'])

    # 创建柱状图 (阶段任务数)
    bar_chart = BarChart()
    bar_chart.title = "各阶段任务数量"
    data = Reference(analysis_sheet, min_col=2, min_row=2, max_row=len(phase_counts) + 1)
    cats = Reference(analysis_sheet, min_col=1, min_row=3, max_row=len(phase_counts) + 1)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    analysis_sheet.add_chart(bar_chart, "A12")

    # 创建饼图 (负责人任务分布)
    pie_chart = PieChart()
    pie_chart.title = "各主责单位/负责人任务分布"
    labels = Reference(analysis_sheet, min_col=4, min_row=3, max_row=len(person_counts) + 2)
    # 数据从E2(包含表头)到数据末尾
    data = Reference(analysis_sheet, min_col=5, min_row=2, max_row=len(person_counts) + 2)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    analysis_sheet.add_chart(pie_chart, "H12")

    # 5. 填充“原始数据”工作表
    # 从DataFrame中移除临时的datetime列
    df_for_sheet = df.drop(columns=['start_date_dt', 'end_date_dt'])
    for r in dataframe_to_rows(df_for_sheet, index=False, header=True):
        data_sheet.append(r)

    # 6. 绘制甘特图
    # 颜色定义
    colors = ['B7DEE8', 'FCD5B4', 'C4D79B', 'D8BFD8', 'F0C2C2', 'FFFACD']
    light_colors = ['E8F5F8', 'FEF3E6', 'EFF5DA', 'F4ECF7', 'FBE9E7', 'FFFDE7']
    unique_phases = df['阶段'].dropna().unique()
    phase_color_map = {phase: {'task': colors[i % len(colors)], 'info': light_colors[i % len(light_colors)]} for
                       i, phase in enumerate(unique_phases)}

    # b. 样式定义
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # c. 【修正】计算日历范围 - 只使用出现过的起止日期
    valid_dates_df = df.dropna(subset=['start_date_dt', 'end_date_dt'])
    calendar_dates = []
    if not valid_dates_df.empty:
        all_dates = set()
        for _, row in valid_dates_df.iterrows():
            all_dates.add(row['start_date_dt'])
            all_dates.add(row['end_date_dt'])
        calendar_dates = sorted(list(all_dates))

    # d. 绘制表头
    headers = ["阶段", "工作子项", "成果文件", "主责单位/负责人"]
    for i, header in enumerate(headers):
        cell = gantt_sheet.cell(row=2, column=i + 1, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, date in enumerate(calendar_dates):
        cell = gantt_sheet.cell(row=2, column=len(headers) + 1 + i, value=date)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = 'YYYY/MM/DD'

    gantt_sheet.merge_cells(start_row=1, start_column=len(headers) + 1, end_row=1,
                            end_column=len(headers) + len(calendar_dates))
    title_cell = gantt_sheet.cell(row=1, column=len(headers) + 1, value="工作日")
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # e. 【修正】按阶段分组并填充数据，以支持合并单元格
    ordered_phases = df['阶段'].dropna().unique()
    current_row = 3  # 数据从第3行开始

    for phase in ordered_phases:
        group = df[df['阶段'] == phase]
        if group.empty:
            continue

        start_row_for_merge = current_row
        end_row_for_merge = current_row + len(group) - 1

        # 合并“阶段”单元格
        if len(group) > 1:
            gantt_sheet.merge_cells(start_row=start_row_for_merge, start_column=1, end_row=end_row_for_merge,
                                    end_column=1)

        colorset = phase_color_map.get(phase, {'info': 'FFFFFF', 'task': 'D9D9D9'})
        info_fill = PatternFill(start_color=colorset['info'], end_color=colorset['info'], fill_type="solid")

        phase_cell = gantt_sheet.cell(row=start_row_for_merge, column=1, value=phase)
        phase_cell.fill = info_fill
        phase_cell.alignment = center_align
        phase_cell.font = Font(bold=True)

        # 循环处理组内的每一行任务
        for i, task in group.iterrows():
            row_to_write = current_row + list(group.index).index(i)

            gantt_sheet.cell(row=row_to_write, column=2, value=task['工作子项']).fill = info_fill
            gantt_sheet.cell(row=row_to_write, column=3, value=task['成果文件']).fill = info_fill
            gantt_sheet.cell(row=row_to_write, column=4, value=task['主责单位/负责人']).fill = info_fill

            if pd.notna(task['start_date_dt']) and pd.notna(task['end_date_dt']):
                task_fill = PatternFill(start_color=colorset['task'], end_color=colorset['task'], fill_type="solid")
                for d_idx, cal_date in enumerate(calendar_dates):
                    if task['start_date_dt'] <= cal_date <= task['end_date_dt']:
                        gantt_sheet.cell(row=row_to_write, column=len(headers) + 1 + d_idx).fill = task_fill

        current_row += len(group)

    # f. 添加边框和冻结窗格
    max_row = current_row - 1
    max_col = len(headers) + len(calendar_dates)
    if max_row > 0 and max_col > 0:
        for row in gantt_sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = thin_border
    gantt_sheet.freeze_panes = 'E3'

    # g. 设置列宽
    gantt_sheet.column_dimensions['A'].width = 15
    gantt_sheet.column_dimensions['B'].width = 30
    gantt_sheet.column_dimensions['C'].width = 25
    gantt_sheet.column_dimensions['D'].width = 20
    for i in range(len(calendar_dates)):
        gantt_sheet.column_dimensions[chr(ord('E') + i)].width = 12

    # --- 修正代码结束 ---

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

